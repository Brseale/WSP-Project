from bs4 import BeautifulSoup
from tqdm import tqdm
import matplotlib.pyplot as plt
import pandas as pd
from collections import Counter
import seaborn as sns
from geopy.geocoders import Nominatim
import geopandas as gpd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
import os

# Create a folder to store the plots
if not os.path.exists('plots'):
    os.makedirs('plots')

# Read in the cover songs
def read_cover_songs(file_path):
    with open(file_path, 'r') as f:
        cover_songs = [line.strip() for line in f.readlines()]
    return cover_songs

# Parse the xml file of all the shows
def parse_xml(xml_file):
    with open(xml_file, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'lxml-xml')

        shows = []

        # Extract show details
        for show in soup.find_all('show'):
            date = show.find('date').get_text()
            location = show.find('location').get_text()
            setlist = [song.get_text() for song in show.find_all('song')]
            shows.append({
                'date': date,
                'location': location,
                'setlist': setlist,
                'num_songs': len(setlist)
            })

    return shows

# create the dataframe
def create_dataframe(shows):
    data = {
        'date': [],
        'location': [],
        'num_songs': [],
        'setlist': []
    }
    for show in shows:
        data['date'].append(show['date'])
        data['location'].append(show['location'])
        data['num_songs'].append(show['num_songs'])
        data['setlist'].append(show['setlist'])

    df = pd.DataFrame(data)
    return df

# Get stats for cover songs
def get_cover_song_stats(df, cover_songs):
    cover_song_counts = Counter()

    normalized_cover_songs = [song.lower().strip() for song in cover_songs]

    for setlist in df['setlist']:
        for song in setlist:
            normalized_song = song.lower().strip()

            if normalized_song in normalized_cover_songs:
                cover_song_counts[song] += 1 

    return cover_song_counts

# Plot functions that save the plots as PNG files
def save_plot_to_file(plot_func, df, filename):
    plot_func(df)
    filepath = os.path.join('plots', filename)
    plt.savefig(filepath)
    plt.close()
    return filepath

# Plot 1: number of songs per show
def plot_num_songs_per_show(df):
    df['date'] = pd.to_datetime(df['date'])
    df_sorted = df.sort_values('date')

    plt.figure(figsize=(10, 6))
    plt.bar(df_sorted['date'], df_sorted['num_songs'])
    plt.xlabel('Date')
    plt.ylabel('Number of Songs')
    plt.title('Number of Songs Per Show')
    plt.xticks(rotation=45)
    plt.tight_layout()
    #plt.show()

# Plot 2: Most frequently played songs
def plot_most_frequent_songs(df):
    # Flatten all setlists into a single list of songs
    all_songs = [song for setlist in df['setlist'] for song in setlist]
    
    # Use Counter to get the frequency of each song
    song_counter = Counter(all_songs)
    
    # Get the top 10 most frequently played songs
    most_common_songs = song_counter.most_common(20)
    
    # Split the most common songs into two lists: names and counts
    song_names, song_counts = zip(*most_common_songs)
    
    plt.figure(figsize=(10, 6))
    plt.bar(song_names, song_counts)
    plt.xlabel('Song')
    plt.ylabel('Number of Times Played')
    plt.title('Top 10 Most Frequently Played Songs')
    plt.xticks(rotation=45)
    plt.tight_layout()
    #plt.show()

# Plot 3: Top 20 locations by number of shows
def plot_song_distribution_across_locations_bar(df, top_n=20):
    # Get the top N locations
    location_counts = df['location'].value_counts().head(top_n)

    # Plot a horizontal bar chart
    plt.figure(figsize=(10, 6))
    location_counts.plot(kind='barh', color='skyblue')
    plt.xlabel('Number of Shows')
    plt.ylabel('Location')
    plt.title(f'Top {top_n} Locations by Number of Shows')
    plt.tight_layout()
    #plt.show()

# Plot 4: Song repetition over time
def plot_song_repetition_over_time(df):
    # Create a list to track whether the song was played at each show
    song_name = 'Disco'
    dates = pd.to_datetime(df['date'])
    song_played = [1 if song_name in setlist else 0 for setlist in df['setlist']]
    
    plt.figure(figsize=(10, 6))
    plt.plot(dates, song_played, marker='o')
    plt.xlabel('Date')
    plt.ylabel(f'{song_name} Played (1 = Yes, 0 = No)')
    plt.title(f'Repetition of "{song_name}" Over Time')
    plt.xticks(rotation=45)
    plt.tight_layout()

# Plot 5: Most popular closing songs
def plot_most_popular_closing_songs(df):
    closing_songs = [setlist[-1] for setlist in df['setlist'] if setlist]  # Get last song from each setlist
    closing_song_counts = Counter(closing_songs).most_common(10)

    song_names, song_counts = zip(*closing_song_counts)

    plt.figure(figsize=(10, 6))
    plt.bar(song_names, song_counts, color='orange')
    plt.xlabel('Song')
    plt.ylabel('Number of Times Closed')
    plt.title('Top 10 Most Popular Closing Songs')
    plt.xticks(rotation=45)
    plt.tight_layout()

# Plot 6: Most popular opening songs
def plot_most_frequent_opening_songs(df):
    opening_songs = [setlist[0] for setlist in df['setlist'] if setlist]  # Get first song from each setlist
    opening_song_counts = Counter(opening_songs).most_common(10)

    song_names, song_counts = zip(*opening_song_counts)

    plt.figure(figsize=(10, 6))
    plt.bar(song_names, song_counts, color='green')
    plt.xlabel('Song')
    plt.ylabel('Number of Times Opened')
    plt.title('Top 10 Most Popular Opening Songs')
    plt.xticks(rotation=45)
    plt.tight_layout()

# Plot 7: Number of songs per show over time
def plot_num_songs_trend_over_time(df):
    df['date'] = pd.to_datetime(df['date'])
    df_sorted = df.sort_values('date')

    plt.figure(figsize=(10, 6))
    plt.plot(df_sorted['date'], df_sorted['num_songs'], marker='o')
    plt.xlabel('Date')
    plt.ylabel('Number of Songs')
    plt.title('Trend of Number of Songs Per Show Over Time')
    plt.xticks(rotation=45)
    plt.tight_layout()

# Plot 8: Number of shows per location heat map
def plot_shows_heatmap(df, top_n=20):
    # Add month column to the DataFrame
    df['month'] = pd.to_datetime(df['date']).dt.month

    # Group by location and month and count shows
    location_month_counts = df.groupby(['location', 'month']).size().unstack(fill_value=0)

    # Limit to top N locations based on total number of shows
    top_locations = location_month_counts.sum(axis=1).nlargest(top_n).index
    location_month_counts = location_month_counts.loc[top_locations]

    # Set up the figure size and adjust aspect ratio
    plt.figure(figsize=(14, 8))

    # Generate heatmap with rotated x-ticks, smaller font for y-axis, and adjusted color scaling
    sns.heatmap(location_month_counts, annot=True, cmap='coolwarm', fmt='d', linewidths=.5, linecolor='white')

    # Rotate x-ticks for months
    plt.xticks(rotation=45)

    # Set title and axis labels
    plt.title('Number of Shows by Location and Month (Top Locations)')
    plt.xlabel('Month')
    plt.ylabel('Location')

    # Tight layout for better spacing
    plt.tight_layout()

# Plot 9: Least frequently played songs
def plot_least_frequent_songs(df):
    num_songs = 20
    all_songs = [song for setlist in df['setlist'] for song in setlist]
    
    song_counter = Counter(all_songs)
    least_common_songs = song_counter.most_common()[:-num_songs-1:-1]
    song_names, song_counts = zip(*least_common_songs) if least_common_songs else ([], [])

    # plot chart
    if song_names and song_counts:
        plt.figure(figsize=(10, 6))
        plt.bar(song_names, song_counts, color='red')
        plt.xlabel('Song')
        plt.ylabel('Number of Times Played')
        plt.title(f'Top {num_songs} Least Frequently Played')
        plt.xticks(rotation=90)
        plt.tight_layout()

# Plot 10: Most popular cover songs
def plot_popular_cover_songs(cover_song_counts, top_n=20):
    most_common_covers = cover_song_counts.most_common(top_n)
    if most_common_covers:
        song_names, play_counts = zip(*most_common_covers)

        plt.figure(figsize=(10, 6))
        plt.bar(song_names, play_counts, color='blue')
        plt.xlabel('Cover Songs')
        plt.ylabel('Number of Times Played')
        plt.title(f'Top {top_n} Most Frequently Played Cover Songs')
        plt.xticks(rotation=90)
        plt.tight_layout()

# Plot 11: Least popular cover song
def plot_least_popular_cover_songs(cover_song_counts):
    num_songs = 20
    least_common_covers = cover_song_counts.most_common()[:-num_songs-1:-1]
    if least_common_covers:
        song_names, play_counts = zip(*least_common_covers)

        plt.figure(figsize=(10, 6))
        plt.bar(song_names, play_counts, color='blue')
        plt.xlabel('Cover Songs')
        plt.ylabel('Number of Times Played')
        plt.title(f'Top {num_songs} Least Frequently Played Cover Songs')
        plt.xticks(rotation=90)
        plt.tight_layout()


def get_location_coordinates(location_name):
    geolocator = Nominatim(user_agent="location_mapper")
    location = geolocator.geocode(location_name)
    if location:
        return location.latitude, location.longitude
    else:
        return None, None

def plot_us_map_with_locations(df):
    # Load the shapefile for the United States from the local path
    world = gpd.read_file('/Users/brooksseale/WSP/ne_110m_admin_0_countries')
    us = world[world['NAME'] == "United States"]

    # Create a new column for coordinates (latitude and longitude)
    df['latitude'] = None
    df['longitude'] = None

    # Convert location names to coordinates
    for idx, row in tqdm(df.iterrows(), desc="10/10: Plotting locations"):
        lat, lon = get_location_coordinates(row['location'])
        df.at[idx, 'latitude'] = lat
        df.at[idx, 'longitude'] = lon

    # Filter rows with valid coordinates (remove None or NaN)
    df = df.dropna(subset=['latitude', 'longitude'])

    # If no valid coordinates are available, raise an error
    if df.empty:
        print("No valid locations to plot.")
        return

    # Plot the U.S. map
    fig, ax = plt.subplots(figsize=(10, 10))
    us.plot(ax=ax, color='whitesmoke', edgecolor='black')

    # Scatter plot for the locations based on latitude and longitude
    sizes = df['num_songs'] * 10  # Scale marker sizes based on the number of shows
    df.plot(kind='scatter', x='longitude', y='latitude', s=sizes, ax=ax, color='blue', alpha=0.6)

    # Explicitly set the aspect ratio to 'equal' to prevent aspect errors
    ax.set_aspect('equal')

    # Set plot title and labels
    plt.title('Locations and Number of Shows Played in the U.S.')
    plt.xlabel('Longitude')
    plt.ylabel('Latitude')

# Create the Excel file and insert images
def export_plots_to_excel(plot_files):
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Plots'

    # Insert each plot into the Excel sheet
    for row, plot_file in enumerate(plot_files, start=1):
        img = ExcelImage(plot_file)
        ws.add_image(img, f'A{row * 15}')  # Adjust positioning of images in rows

    # Save the workbook
    wb.save('show_plots.xlsx')

# Create Excel file with all of the show data
def create_excel_with_show_data(df):
    all_rows = []

    for index, row in df.iterrows():
        for song in row['setlist']:
            all_rows.append({
                'Location': row['location'],
                'Date': row['date'],
                'Song': song
            })

    # Song DF
    song_df = pd.DataFrame(all_rows)

    # Count num times song played
    song_counter = Counter(song_df['Song'])
    song_df["Times Played"] = song_df['Song'].apply(lambda x: song_counter[x])

    # Sort df by song freq
    song_df_sorted = song_df.sort_values(by='Times Played', ascending=False)

    # Create xl
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Show Data'

    # Write df to excel
    for row in dataframe_to_rows(song_df_sorted, index=False, header=False):
        ws.append(row)

    wb.save('all_show_data.xlsx')

def get_last_three_show_dates(df):
    last_three_show_dates = []
    count = 0

    for index, row in df.iterrows():
        if count < 3:  # Add 3 dates, not 2
            last_three_show_dates.append(row['date'])
            count += 1
        else:
            break  # Exit loop when 3 dates are added
    
    return last_three_show_dates

def get_date_code(curr_date, last_three_show_dates):
    for i in range(3):
        if curr_date in last_three_show_dates[i]:
            return i + 1
        
    return 0

# Create Excel file with all cover song data
def create_excel_with_cover_songs(df, cover_songs):
    all_rows = []
    last_three_show_dates = get_last_three_show_dates(df)

    normalized_cover_songs = [song.lower().strip() for song in cover_songs]

    for index, row in df.iterrows():
        for song in row['setlist']:
            normalized_song = song.lower().strip()

            if normalized_song in normalized_cover_songs:
                all_rows.append({
                    'Location': row['location'],
                    'Date': row['date'],
                    'Song': song,
                    'Song Type': 'Cover',
                    'Recently Played': get_date_code(row['date'], last_three_show_dates)
                })
            else:
                all_rows.append({
                    'Location': row['location'],
                    'Date': row['date'],
                    'Song': song,
                    'Song Type': 'Original',
                    'Recently Played': get_date_code(row['date'], last_three_show_dates)
                })

    cover_song_df = pd.DataFrame(all_rows)
    cover_song_df['Date'] = pd.to_datetime(cover_song_df['Date'], errors='coerce')

    # Count how many times each song was played
    song_counter = Counter(cover_song_df['Song'])
    cover_song_df["Times Played"] = cover_song_df['Song'].apply(lambda x: song_counter[x])

    # Sort the DataFrame by the frequency of times played
    cover_song_df_sorted = cover_song_df.sort_values(by='Times Played', ascending=False)

    # Create the Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cover Songs Data'

    # Write the DataFrame to Excel
    for row in dataframe_to_rows(cover_song_df_sorted, index=False, header=True):
        ws.append(row)

    # Save the Excel workbook
    wb.save('all_songs_data.xlsx')

def get_html_from_excel_table(file, excel_sheet_name):
    df = pd.read_excel(file, sheet_name=excel_sheet_name)
    df.to_html('song_table.html', index=False)


if __name__ == "__main__":

    # Read in cover songs
    cover_songs_file = 'txt_files/all_covers.txt'
    cover_songs = read_cover_songs(cover_songs_file)

    # Parse the XML file
    xml_file = 'xml_files/allshows_setlistfm.xml'
    shows = parse_xml(xml_file)

    # Create a DataFrame from the shows
    df = create_dataframe(shows)

    # create_excel_with_cover_songs(df, cover_songs)

    excel_file = 'excel_files/WSP_All_Show_Data.xlsx'
    sheet_name = 'All_Song_Data'
    get_html_from_excel_table(excel_file, sheet_name)

    #create_excel_with_show_data(df)

    # Save plots to PNG files
    # plot_files = []
    # plot_files.append(save_plot_to_file(plot_num_songs_per_show, df, 'plot_num_songs_per_show.png'))
    # plot_files.append(save_plot_to_file(plot_most_frequent_songs, df, 'plot_most_frequent_songs.png'))
    # plot_files.append(save_plot_to_file(plot_song_distribution_across_locations_bar, df, 'plot_song_distribution_across_locations_bar.png'))
    # plot_files.append(save_plot_to_file(plot_song_repetition_over_time, df, 'plot_song_repetition_over_time.png'))
    # plot_files.append(save_plot_to_file(plot_most_popular_closing_songs, df, 'plot_most_popular_closing_songs.png'))
    # plot_files.append(save_plot_to_file(plot_most_frequent_opening_songs, df, 'plot_most_frequent_opening_songs.png'))
    # plot_files.append(save_plot_to_file(plot_num_songs_trend_over_time, df, 'plot_num_songs_trend_over_time.png'))
    # plot_files.append(save_plot_to_file(plot_shows_heatmap, df, 'plot_shows_heatmap.png'))

    # Export plots to Excel
    #export_plots_to_excel(plot_files)

    # Plot 1
    plot_num_songs_per_show(df)

    # Plot 2
    plot_most_frequent_songs(df)

    # Plot 3
    plot_song_distribution_across_locations_bar(df, top_n=20)

    # Plot 4
    plot_song_repetition_over_time(df)

    # Plot 5
    plot_most_popular_closing_songs(df)

    # Plot 6
    plot_most_frequent_opening_songs(df)

    # Plot 7
    plot_num_songs_trend_over_time(df)

    # Plot 8
    plot_shows_heatmap(df, top_n=20)

    # Plot 9
    plot_least_frequent_songs(df)

    # Plot 10 
    cover_song_counts = get_cover_song_stats(df, cover_songs)
    plot_popular_cover_songs(cover_song_counts, top_n=20)

    # Plot 11
    plot_least_popular_cover_songs(cover_song_counts)

    #plot_us_map_with_locations(df)

    plt.show()

    #/Users/brooksseale/Desktop/WSP/ne_110m_admin_0_countries/ne_110m_admin_0_countries.shp